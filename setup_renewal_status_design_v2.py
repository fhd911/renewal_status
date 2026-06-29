from pathlib import Path
import re

ROOT = Path.cwd()


def write_file(path: str, content: str):
    p = ROOT / path
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(content.strip() + "\n", encoding="utf-8")
    print(f"✓ كتب: {path}")


def patch_settings():
    settings = ROOT / "config" / "settings.py"
    if not settings.exists():
        raise SystemExit("لم أجد config/settings.py — شغّل الملف من داخل مجلد مشروع Django نفسه.")
    s = settings.read_text(encoding="utf-8")

    if '"inquiry"' not in s and "'inquiry'" not in s:
        s = re.sub(
            r"INSTALLED_APPS\s*=\s*\[(.*?)\]",
            lambda m: "INSTALLED_APPS = [" + m.group(1) + "\n    'inquiry',\n]",
            s,
            flags=re.S,
        )

    # ضمان اللغة والوقت المناسبين
    s = re.sub(r"LANGUAGE_CODE\s*=\s*['\"].*?['\"]", "LANGUAGE_CODE = 'ar'", s)
    s = re.sub(r"TIME_ZONE\s*=\s*['\"].*?['\"]", "TIME_ZONE = 'Asia/Riyadh'", s)

    # لا نحتاج STATICFILES_DIRS لأننا نستخدم static داخل التطبيق inquiry/static
    settings.write_text(s, encoding="utf-8")
    print("✓ تم تحديث config/settings.py")


def patch_urls():
    urls = ROOT / "config" / "urls.py"
    if not urls.exists():
        raise SystemExit("لم أجد config/urls.py")
    content = '''
from django.contrib import admin
from django.urls import path, include

urlpatterns = [
    path("admin/", admin.site.urls),
    path("", include("inquiry.urls")),
]
'''
    urls.write_text(content.strip() + "\n", encoding="utf-8")
    print("✓ تم تحديث config/urls.py")


models_py = r'''
from django.db import models


class RenewalRecord(models.Model):
    national_id = models.CharField("السجل المدني", max_length=20, db_index=True)
    full_name = models.CharField("الاسم", max_length=255)
    school_name = models.CharField("المدرسة", max_length=255, blank=True)
    sector = models.CharField("القطاع التعليمي", max_length=255, blank=True)
    category = models.CharField("الفئة", max_length=120, blank=True)
    raw_status = models.CharField("الحالة الأصلية", max_length=255, blank=True)
    status_key = models.CharField("رمز الحالة", max_length=60, blank=True)
    status_label = models.CharField("الحالة المعروضة", max_length=120, blank=True)
    status_message = models.TextField("رسالة الحالة", blank=True)
    is_published = models.BooleanField("منشور", default=True)
    created_at = models.DateTimeField(auto_now_add=True)
    updated_at = models.DateTimeField(auto_now=True)

    class Meta:
        verbose_name = "سجل استعلام"
        verbose_name_plural = "سجلات الاستعلام"
        ordering = ["full_name"]

    def __str__(self):
        return f"{self.full_name} - {self.national_id}"


class InquiryLog(models.Model):
    national_id = models.CharField("السجل المدني", max_length=20, db_index=True)
    found = models.BooleanField("وجد نتيجة", default=False)
    ip_address = models.GenericIPAddressField("عنوان IP", null=True, blank=True)
    user_agent = models.TextField("المتصفح", blank=True)
    created_at = models.DateTimeField(auto_now_add=True)

    class Meta:
        verbose_name = "سجل دخول"
        verbose_name_plural = "سجلات الدخول"
        ordering = ["-created_at"]

    def __str__(self):
        return f"{self.national_id} - {'وجد' if self.found else 'غير موجود'}"
'''

admin_py = r'''
from django.contrib import admin
from .models import RenewalRecord, InquiryLog


@admin.register(RenewalRecord)
class RenewalRecordAdmin(admin.ModelAdmin):
    list_display = ("full_name", "national_id", "category", "school_name", "sector", "status_label", "is_published")
    list_filter = ("category", "sector", "status_label", "is_published")
    search_fields = ("full_name", "national_id", "school_name", "sector")
    list_per_page = 50


@admin.register(InquiryLog)
class InquiryLogAdmin(admin.ModelAdmin):
    list_display = ("national_id", "found", "ip_address", "created_at")
    list_filter = ("found", "created_at")
    search_fields = ("national_id", "ip_address")
    list_per_page = 50
'''

utils_py = r'''
import re

ARABIC_DIGITS = str.maketrans("٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹", "01234567890123456789")


def normalize_national_id(value):
    if value is None:
        return ""
    text = str(value).strip().translate(ARABIC_DIGITS)
    if text.endswith(".0"):
        text = text[:-2]
    return re.sub(r"\D+", "", text)


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def classify_status(raw_status):
    text = clean_text(raw_status)
    compact = text.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا").replace("ة", "ه")

    if any(word in compact for word in ["عدم", "لا يجدد", "لم يجدد", "غير مجدد"]):
        return {
            "key": "not_renewed",
            "label": "لم يتم التجديد",
            "message": "وفق ما ورد في البيانات المعتمدة لدى قسم الإدارة المدرسية.",
        }

    if "يجدد الى" in compact or "يجدد الي" in compact or "لمده" in compact or "مدة" in text:
        return {
            "key": "limited",
            "label": "تم التجديد لمدة محددة",
            "message": "تم التجديد وفق المدة المحددة في البيانات المعتمدة.",
        }

    if any(word in compact for word in ["رفع", "صاحب الصلاحيه", "تحت الاجراء", "تحت الدراسه"]):
        return {
            "key": "pending",
            "label": "تحت الإجراء",
            "message": "الحالة مرفوعة أو تحت الإجراء حسب البيانات المعتمدة.",
        }

    if any(word in compact for word in ["تجديد", "مجدد", "تم التجديد", "يرشح", "يكلف"]):
        return {
            "key": "renewed",
            "label": "تم التجديد",
            "message": "نبارك لكم تجديد التكليف، سائلين الله لكم التوفيق والسداد.",
        }

    return {
        "key": "info",
        "label": text or "غير متوفر",
        "message": "تظهر الحالة وفق البيانات المعتمدة لدى قسم الإدارة المدرسية.",
    }
'''

views_py = r'''
from django.shortcuts import render
from .models import RenewalRecord, InquiryLog
from .utils import normalize_national_id


def get_client_ip(request):
    x_forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR")
    if x_forwarded_for:
        return x_forwarded_for.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR")


def home(request):
    context = {
        "searched": False,
        "national_id": "",
        "records": [],
        "error": "",
    }

    if request.method == "POST":
        national_id = normalize_national_id(request.POST.get("national_id"))
        context["searched"] = True
        context["national_id"] = national_id

        if not national_id:
            context["error"] = "فضلاً أدخل السجل المدني." 
            return render(request, "inquiry/home.html", context)

        records = list(
            RenewalRecord.objects.filter(
                national_id=national_id,
                is_published=True,
            ).order_by("category", "school_name")
        )

        InquiryLog.objects.create(
            national_id=national_id,
            found=bool(records),
            ip_address=get_client_ip(request),
            user_agent=request.META.get("HTTP_USER_AGENT", "")[:1000],
        )

        if records:
            context["records"] = records
        else:
            context["error"] = "لا توجد نتيجة مطابقة للسجل المدني المدخل."

    return render(request, "inquiry/home.html", context)
'''

urls_py = r'''
from django.urls import path
from . import views

app_name = "inquiry"

urlpatterns = [
    path("", views.home, name="home"),
]
'''

command_py = r'''
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from inquiry.models import RenewalRecord
from inquiry.utils import normalize_national_id, clean_text, classify_status


ALIASES = {
    "full_name": ["الاسم الرباعي", "الاسم", "اسم الموظف", "اسم المستفيد", "الموظف"],
    "national_id": ["السجل المدني", "رقم الهوية", "الهوية الوطنية", "السجل", "سجل مدني"],
    "school_name": ["اسم المدرسة", "المدرسة", "المدرسة الحالية"],
    "sector": ["القطاع التعليمي", "القطاع", "قطاع"],
    "category": ["الفئة", "العمل الحالي", "المسمى", "نوع التكليف", "الوظيفة"],
    "status": ["الحالة", "الحلة", "حالة التجديد", "التوصية", "نوع التوصية"],
}


def normalize_header(value):
    return clean_text(value).replace("ـ", "").replace(" ", "").lower()


def find_col(headers, aliases):
    normalized = {normalize_header(h): i for i, h in enumerate(headers)}
    for alias in aliases:
        key = normalize_header(alias)
        if key in normalized:
            return normalized[key]
    # بحث مرن داخل العنوان
    for i, h in enumerate(headers):
        nh = normalize_header(h)
        for alias in aliases:
            if normalize_header(alias) in nh:
                return i
    return None


class Command(BaseCommand):
    help = "استيراد بيانات حالة التجديد من ملف إكسل"

    def add_arguments(self, parser):
        parser.add_argument("excel_path", type=str)
        parser.add_argument("--clear", action="store_true", help="حذف السجلات السابقة قبل الاستيراد")

    def handle(self, *args, **options):
        excel_path = options["excel_path"]

        try:
            wb = load_workbook(excel_path, data_only=True)
        except Exception as exc:
            raise CommandError(f"تعذر فتح ملف الإكسل: {exc}")

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise CommandError("ملف الإكسل فارغ.")

        header_row_index = None
        headers = None
        for idx, row in enumerate(rows[:10]):
            values = [clean_text(v) for v in row]
            joined = " ".join(values)
            if "السجل" in joined or "الهوية" in joined:
                header_row_index = idx
                headers = values
                break

        if headers is None:
            raise CommandError("لم أتمكن من تحديد صف العناوين في ملف الإكسل.")

        cols = {key: find_col(headers, aliases) for key, aliases in ALIASES.items()}
        required = ["full_name", "national_id", "status"]
        missing = [key for key in required if cols.get(key) is None]
        if missing:
            raise CommandError(f"أعمدة مطلوبة غير موجودة: {', '.join(missing)}")

        if options["clear"]:
            RenewalRecord.objects.all().delete()
            self.stdout.write(self.style.WARNING("تم حذف السجلات السابقة."))

        created = 0
        skipped = 0

        for row in rows[header_row_index + 1:]:
            if not row or not any(row):
                continue

            national_id = normalize_national_id(row[cols["national_id"]] if cols["national_id"] is not None else "")
            full_name = clean_text(row[cols["full_name"]] if cols["full_name"] is not None else "")
            raw_status = clean_text(row[cols["status"]] if cols["status"] is not None else "")

            if not national_id or not full_name:
                skipped += 1
                continue

            status = classify_status(raw_status)

            RenewalRecord.objects.create(
                national_id=national_id,
                full_name=full_name,
                school_name=clean_text(row[cols["school_name"]]) if cols.get("school_name") is not None else "",
                sector=clean_text(row[cols["sector"]]) if cols.get("sector") is not None else "",
                category=clean_text(row[cols["category"]]) if cols.get("category") is not None else "",
                raw_status=raw_status,
                status_key=status["key"],
                status_label=status["label"],
                status_message=status["message"],
                is_published=True,
            )
            created += 1

        self.stdout.write(self.style.SUCCESS(f"تم الاستيراد بنجاح: {created} سجل"))
        if skipped:
            self.stdout.write(self.style.WARNING(f"تم تجاوز {skipped} صف بسبب نقص الاسم أو السجل المدني."))
'''

home_html = r'''
{% load static %}
<!doctype html>
<html lang="ar" dir="rtl">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>استعلام حالة التكليف</title>
  <link rel="stylesheet" href="{% static 'inquiry/style.css' %}">
</head>
<body>
  <main class="page-shell">
    <section class="hero-card">
      <div class="brand-row">
        <div class="emblem">
          <span></span>
        </div>
        <div>
          <p class="eyebrow">الإدارة العامة للتعليم بمنطقة عسير</p>
          <h1>استعلام حالة التكليف</h1>
          <p class="subtitle">خدمة إلكترونية للاستعلام عن حالة التجديد لفئات التشكيلات المدرسية.</p>
        </div>
      </div>

      <form method="post" class="search-box" autocomplete="off">
        {% csrf_token %}
        <label for="national_id">السجل المدني</label>
        <div class="input-action">
          <input id="national_id" name="national_id" value="{{ national_id }}" inputmode="numeric" maxlength="10" placeholder="أدخل السجل المدني" autofocus>
          <button type="submit">استعلام</button>
        </div>
        <p class="hint">تظهر النتيجة وفق البيانات المعتمدة لدى قسم الإدارة المدرسية.</p>
      </form>
    </section>

    {% if searched %}
      {% if error %}
        <section class="result-card empty-card">
          <div class="empty-icon">!</div>
          <h2>{{ error }}</h2>
          <p>يرجى التأكد من إدخال السجل المدني بالشكل الصحيح.</p>
        </section>
      {% endif %}

      {% if records %}
        <section class="results-wrap">
          {% for record in records %}
            <article class="result-card status-{{ record.status_key|default:'info' }}">
              <div class="status-head">
                <div>
                  <p class="small-title">حالة التكليف</p>
                  <h2>{{ record.status_label }}</h2>
                </div>
                <div class="status-badge">{{ record.category|default:"التشكيلات المدرسية" }}</div>
              </div>

              <p class="status-message">{{ record.status_message }}</p>

              <div class="details-grid">
                <div class="detail-item">
                  <span>الاسم</span>
                  <strong>{{ record.full_name }}</strong>
                </div>
                <div class="detail-item">
                  <span>الفئة</span>
                  <strong>{{ record.category|default:"غير متوفر" }}</strong>
                </div>
                <div class="detail-item">
                  <span>المدرسة</span>
                  <strong>{{ record.school_name|default:"غير متوفر" }}</strong>
                </div>
                <div class="detail-item">
                  <span>القطاع التعليمي</span>
                  <strong>{{ record.sector|default:"غير متوفر" }}</strong>
                </div>
                <div class="detail-item wide">
                  <span>الحالة في البيانات</span>
                  <strong>{{ record.raw_status|default:record.status_label }}</strong>
                </div>
              </div>
            </article>
          {% endfor %}
        </section>
      {% endif %}
    {% endif %}

    <footer class="footer-note">
      هذه الخدمة مخصصة للاستعلام فقط، ولا تُعد بديلًا عن القرارات أو الخطابات الرسمية المعتمدة.
    </footer>
  </main>
</body>
</html>
'''

style_css = r'''
:root{
  --bg:#eef4f6;
  --card:#ffffff;
  --ink:#17313a;
  --muted:#6d7e86;
  --line:#dce8ec;
  --primary:#146b78;
  --primary-2:#0f515c;
  --gold:#b78945;
  --shadow:0 24px 70px rgba(20,57,70,.13);
  --radius:32px;
}
*{box-sizing:border-box}
body{
  margin:0;
  min-height:100vh;
  font-family:"Cairo","Tajawal","Segoe UI",Tahoma,Arial,sans-serif;
  color:var(--ink);
  background:
    radial-gradient(circle at top right, rgba(20,107,120,.16), transparent 32rem),
    radial-gradient(circle at bottom left, rgba(183,137,69,.13), transparent 28rem),
    linear-gradient(135deg,#f8fbfc 0%,var(--bg) 100%);
}
body::before{
  content:"";
  position:fixed;
  inset:0;
  pointer-events:none;
  background-image:
    linear-gradient(rgba(20,107,120,.045) 1px, transparent 1px),
    linear-gradient(90deg, rgba(20,107,120,.045) 1px, transparent 1px);
  background-size:42px 42px;
  mask-image:linear-gradient(to bottom, black, transparent 80%);
}
.page-shell{
  width:min(1060px, calc(100% - 32px));
  margin:0 auto;
  padding:56px 0 34px;
  position:relative;
}
.hero-card,.result-card{
  background:rgba(255,255,255,.86);
  border:1px solid rgba(220,232,236,.95);
  box-shadow:var(--shadow);
  backdrop-filter:blur(18px);
}
.hero-card{
  border-radius:var(--radius);
  padding:34px;
  overflow:hidden;
  position:relative;
}
.hero-card::after{
  content:"";
  position:absolute;
  width:310px;height:310px;
  left:-110px;top:-120px;
  border-radius:999px;
  background:linear-gradient(135deg, rgba(20,107,120,.13), rgba(183,137,69,.12));
}
.brand-row{
  display:flex;
  gap:18px;
  align-items:center;
  position:relative;
  z-index:1;
}
.emblem{
  width:72px;height:72px;
  border-radius:24px;
  display:grid;
  place-items:center;
  flex:0 0 auto;
  background:linear-gradient(145deg,var(--primary),var(--primary-2));
  box-shadow:0 16px 35px rgba(20,107,120,.28);
}
.emblem span{
  width:32px;height:32px;
  border-radius:12px;
  border:7px solid rgba(255,255,255,.95);
  border-top-color:rgba(183,137,69,.95);
  transform:rotate(45deg);
}
.eyebrow{
  margin:0 0 4px;
  color:var(--gold);
  font-size:15px;
  font-weight:800;
  letter-spacing:.2px;
}
h1{
  margin:0;
  font-size:clamp(30px, 5vw, 54px);
  line-height:1.15;
  font-weight:900;
  color:var(--ink);
}
.subtitle{
  margin:12px 0 0;
  color:var(--muted);
  font-size:17px;
  line-height:1.9;
}
.search-box{
  margin-top:34px;
  position:relative;
  z-index:1;
  background:linear-gradient(180deg,#fff,rgba(255,255,255,.72));
  border:1px solid var(--line);
  border-radius:26px;
  padding:22px;
}
.search-box label{
  display:block;
  margin:0 0 12px;
  font-weight:900;
  font-size:15px;
  color:#2f4a52;
}
.input-action{
  display:flex;
  gap:12px;
}
input{
  width:100%;
  height:62px;
  border:1px solid #d2e2e7;
  border-radius:20px;
  padding:0 20px;
  font:inherit;
  font-size:22px;
  font-weight:800;
  color:var(--ink);
  background:#f9fcfd;
  outline:none;
  transition:.2s ease;
  direction:ltr;
  text-align:center;
}
input:focus{
  border-color:rgba(20,107,120,.65);
  box-shadow:0 0 0 5px rgba(20,107,120,.09);
  background:white;
}
button{
  height:62px;
  min-width:160px;
  border:0;
  border-radius:20px;
  font:inherit;
  font-size:17px;
  font-weight:900;
  color:white;
  cursor:pointer;
  background:linear-gradient(135deg,var(--primary),var(--primary-2));
  box-shadow:0 16px 30px rgba(20,107,120,.22);
  transition:.2s ease;
}
button:hover{transform:translateY(-1px);box-shadow:0 20px 36px rgba(20,107,120,.28)}
.hint{
  margin:12px 4px 0;
  color:var(--muted);
  font-size:13px;
}
.results-wrap{margin-top:22px;display:grid;gap:18px}
.result-card{
  border-radius:28px;
  padding:28px;
  position:relative;
  overflow:hidden;
}
.result-card::before{
  content:"";
  position:absolute;
  inset:0 auto 0 0;
  width:8px;
  background:var(--primary);
}
.status-head{
  display:flex;
  justify-content:space-between;
  gap:20px;
  align-items:flex-start;
}
.small-title{
  margin:0 0 4px;
  color:var(--muted);
  font-size:14px;
  font-weight:800;
}
h2{
  margin:0;
  font-size:clamp(27px,4vw,42px);
  line-height:1.2;
  font-weight:900;
}
.status-badge{
  padding:10px 16px;
  border-radius:999px;
  background:#f2f7f8;
  color:#31545d;
  border:1px solid var(--line);
  font-size:14px;
  font-weight:900;
  white-space:nowrap;
}
.status-message{
  margin:18px 0 22px;
  color:#49656d;
  font-size:17px;
  line-height:1.9;
}
.details-grid{
  display:grid;
  grid-template-columns:repeat(2,minmax(0,1fr));
  gap:12px;
}
.detail-item{
  background:#f8fbfc;
  border:1px solid #e3edf0;
  border-radius:18px;
  padding:15px 16px;
}
.detail-item span{
  display:block;
  color:#7a8d94;
  font-size:12.5px;
  font-weight:800;
  margin-bottom:7px;
}
.detail-item strong{
  display:block;
  color:#17313a;
  font-size:16px;
  line-height:1.65;
}
.detail-item.wide{grid-column:1/-1}
.status-renewed::before{background:#1f8f70}.status-renewed h2{color:#16765d}
.status-not_renewed::before{background:#b64b52}.status-not_renewed h2{color:#a33c44}
.status-pending::before{background:#b78945}.status-pending h2{color:#9a6e2d}
.status-limited::before{background:#2f6fb2}.status-limited h2{color:#225d9b}
.empty-card{margin-top:22px;text-align:center}
.empty-card::before{background:#b78945}
.empty-icon{
  width:54px;height:54px;
  margin:0 auto 12px;
  display:grid;place-items:center;
  border-radius:18px;
  background:#fff4df;
  color:#9a6e2d;
  font-weight:900;
  font-size:26px;
}
.empty-card p{margin:10px 0 0;color:var(--muted)}
.footer-note{
  text-align:center;
  margin-top:24px;
  color:#789098;
  font-size:13px;
  line-height:1.8;
}
@media(max-width:720px){
  .page-shell{width:min(100% - 22px, 1060px);padding-top:22px}
  .hero-card{padding:22px;border-radius:24px}
  .brand-row{align-items:flex-start}
  .emblem{width:58px;height:58px;border-radius:19px}
  .input-action{flex-direction:column}
  button{width:100%}
  .details-grid{grid-template-columns:1fr}
  .status-head{flex-direction:column}
  .status-badge{white-space:normal}
}
'''

# تنفيذ الكتابة والتحديث
patch_settings()
patch_urls()
write_file("inquiry/models.py", models_py)
write_file("inquiry/admin.py", admin_py)
write_file("inquiry/utils.py", utils_py)
write_file("inquiry/views.py", views_py)
write_file("inquiry/urls.py", urls_py)
write_file("inquiry/management/__init__.py", "")
write_file("inquiry/management/commands/__init__.py", "")
write_file("inquiry/management/commands/import_renewal_excel.py", command_py)
write_file("inquiry/templates/inquiry/home.html", home_html)
write_file("inquiry/static/inquiry/style.css", style_css)
(ROOT / "uploads").mkdir(exist_ok=True)
print("\nتم تجهيز التصميم والمنصة بنجاح.")
print("الأوامر التالية:")
print("python manage.py makemigrations inquiry")
print("python manage.py migrate")
print('python manage.py import_renewal_excel ".\\uploads\\استعلام.xlsx" --clear')
print("python manage.py runserver")
