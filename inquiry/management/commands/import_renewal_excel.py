from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from inquiry.models import RenewalRecord
from inquiry.utils import normalize_national_id, clean_text, classify_status


ALIASES = {
    "full_name": ["الاسم الرباعي", "الاسم", "اسم الموظف", "اسم المستفيد", "الموظف"],
    "national_id": ["السجل المدني", "رقم الهوية", "الهوية الوطنية", "السجل", "سجل مدني"],
    "school_name": ["اسم المدرسة", "المدرسة", "المدرسة الحالية"],
    "sector": ["القطاع التعليمي", "القطاع", "قطاع"],
    "category": ["الفئة", "العمل الحالي", "المسمى", "نوع التكليف", "الوظيفة"],
    "status": ["الحالة", "الحلة", "حالة التجديد", "التوصية", "نوع التوصية"],
}


def normalize_header(value):
    return clean_text(value).replace("ـ", "").replace(" ", "").lower()


def find_col(headers, aliases):
    normalized = {normalize_header(h): i for i, h in enumerate(headers)}
    for alias in aliases:
        key = normalize_header(alias)
        if key in normalized:
            return normalized[key]
    # بحث مرن داخل العنوان
    for i, h in enumerate(headers):
        nh = normalize_header(h)
        for alias in aliases:
            if normalize_header(alias) in nh:
                return i
    return None


class Command(BaseCommand):
    help = "استيراد بيانات حالة التجديد من ملف إكسل"

    def add_arguments(self, parser):
        parser.add_argument("excel_path", type=str)
        parser.add_argument("--clear", action="store_true", help="حذف السجلات السابقة قبل الاستيراد")

    def handle(self, *args, **options):
        excel_path = options["excel_path"]

        try:
            wb = load_workbook(excel_path, data_only=True)
        except Exception as exc:
            raise CommandError(f"تعذر فتح ملف الإكسل: {exc}")

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise CommandError("ملف الإكسل فارغ.")

        header_row_index = None
        headers = None
        for idx, row in enumerate(rows[:10]):
            values = [clean_text(v) for v in row]
            joined = " ".join(values)
            if "السجل" in joined or "الهوية" in joined:
                header_row_index = idx
                headers = values
                break

        if headers is None:
            raise CommandError("لم أتمكن من تحديد صف العناوين في ملف الإكسل.")

        cols = {key: find_col(headers, aliases) for key, aliases in ALIASES.items()}
        required = ["full_name", "national_id", "status"]
        missing = [key for key in required if cols.get(key) is None]
        if missing:
            raise CommandError(f"أعمدة مطلوبة غير موجودة: {', '.join(missing)}")

        if options["clear"]:
            RenewalRecord.objects.all().delete()
            self.stdout.write(self.style.WARNING("تم حذف السجلات السابقة."))

        created = 0
        skipped = 0

        for row in rows[header_row_index + 1:]:
            if not row or not any(row):
                continue

            national_id = normalize_national_id(row[cols["national_id"]] if cols["national_id"] is not None else "")
            full_name = clean_text(row[cols["full_name"]] if cols["full_name"] is not None else "")
            raw_status = clean_text(row[cols["status"]] if cols["status"] is not None else "")

            if not national_id or not full_name:
                skipped += 1
                continue

            status = classify_status(raw_status)

            RenewalRecord.objects.create(
                national_id=national_id,
                full_name=full_name,
                school_name=clean_text(row[cols["school_name"]]) if cols.get("school_name") is not None else "",
                sector=clean_text(row[cols["sector"]]) if cols.get("sector") is not None else "",
                category=clean_text(row[cols["category"]]) if cols.get("category") is not None else "",
                raw_status=raw_status,
                status_key=status["key"],
                status_label=status["label"],
                status_message=status["message"],
                is_published=True,
            )
            created += 1

        self.stdout.write(self.style.SUCCESS(f"تم الاستيراد بنجاح: {created} سجل"))
        if skipped:
            self.stdout.write(self.style.WARNING(f"تم تجاوز {skipped} صف بسبب نقص الاسم أو السجل المدني."))
