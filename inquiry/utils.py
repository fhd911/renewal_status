import re
from collections import Counter

from openpyxl import load_workbook


CATEGORY_MANAGER = "مدير مدرسة"
CATEGORY_DEPUTY = "وكيل مدرسة"
CATEGORY_COUNSELOR = "موجه طلابي"
CATEGORY_ACTIVITY = "رائد نشاط"
CATEGORY_UNKNOWN = "غير متوفر"

CATEGORY_GROUPS = {
    "manager": CATEGORY_MANAGER,
    "deputy": CATEGORY_DEPUTY,
    "counselor": CATEGORY_COUNSELOR,
    "activity_leader": CATEGORY_ACTIVITY,
    "unknown": CATEGORY_UNKNOWN,
}

STATUS_LABELS = {
    "renewed": "تم التجديد",
    "not_renewed": "لم يتم التجديد",
    "limited": "تم التجديد لمدة محددة",
    "pending": "تحت الإجراء",
}


def normalize_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_national_id(value):
    text = normalize_text(value)
    if text.endswith(".0"):
        text = text[:-2]
    return re.sub(r"\D", "", text)


def normalize_mobile_last4(value):
    text = normalize_text(value)
    if text.endswith(".0"):
        text = text[:-2]
    digits = re.sub(r"\D", "", text)
    if not digits:
        return ""
    return digits[-4:]


def normalize_ministry_number(value):
    text = normalize_text(value)
    if text.endswith(".0"):
        text = text[:-2]
    return text


def normalize_category(value):
    text = normalize_text(value)
    if not text:
        return CATEGORY_UNKNOWN

    compact = text.replace("ـ", "").replace("/", " ").strip()

    if any(word in compact for word in ["موجه", "موجهة", "مرشد", "مرشدة", "طلابي", "طلابية"]):
        return CATEGORY_COUNSELOR

    if any(word in compact for word in ["رائد", "رائدة", "نشاط"]):
        return CATEGORY_ACTIVITY

    if any(word in compact for word in ["وكيل", "وكيلة"]):
        return CATEGORY_DEPUTY

    if any(word in compact for word in ["مدير", "مديرة"]):
        return CATEGORY_MANAGER

    return text


def category_group_key(category):
    category = normalize_category(category)

    if category == CATEGORY_MANAGER:
        return "manager"
    if category == CATEGORY_DEPUTY:
        return "deputy"
    if category == CATEGORY_COUNSELOR:
        return "counselor"
    if category == CATEGORY_ACTIVITY:
        return "activity_leader"

    return "unknown"


def classify_status(value):
    text = normalize_text(value)
    compact = text.replace("ـ", "").replace("إ", "ا").replace("أ", "ا").replace("آ", "ا")

    if not compact:
        return {
            "key": "pending",
            "label": "تحت الإجراء",
            "message": "لم تظهر حالة نهائية في البيانات المعتمدة.",
        }

    if any(word in compact for word in ["عدم تجديد", "لم يتم التجديد", "لا يجدد", "لايجدد", "غير مجدد", "انهاء التكليف", "إنهاء التكليف"]):
        return {
            "key": "not_renewed",
            "label": "لم يتم التجديد",
            "message": "تظهر حالة التكليف في البيانات المعتمدة بأنه لم يتم التجديد، ويمكن مراجعة جهة الاختصاص للاستفسار وفق القنوات الرسمية.",
        }

    if any(word in compact for word in ["يجدد الى", "يجدد إلى", "مدة محددة", "لمدة محددة", "الى نهاية", "إلى نهاية"]):
        return {
            "key": "limited",
            "label": "تم التجديد لمدة محددة",
            "message": "تم التجديد وفق المدة المحددة في البيانات المعتمدة.",
        }

    if any(word in compact for word in ["رفع", "صاحب الصلاحية", "تحت الاجراء", "تحت الإجراء", "بانتظار", "قيد"]):
        return {
            "key": "pending",
            "label": "تحت الإجراء",
            "message": "حالة التكليف ما زالت تحت الإجراء وفق البيانات المعتمدة.",
        }

    if any(word in compact for word in ["تم التجديد", "تجديد تكليف", "تجديد", "مجدد", "يجدد"]):
        return {
            "key": "renewed",
            "label": "تم التجديد",
            "message": "نبارك لكم تجديد التكليف، سائلين الله لكم التوفيق والسداد.",
        }

    return {
        "key": "pending",
        "label": text,
        "message": "تظهر الحالة كما وردت في البيانات المعتمدة.",
    }


HEADER_ALIASES = {
    "national_id": [
        "السجل المدني", "سجل مدني", "رقم السجل المدني", "الهوية", "رقم الهوية",
        "الهوية الوطنية", "رقم الهوية الوطنية", "السجل"
    ],
    "full_name": [
        "الاسم الرباعي", "الاسم", "اسم الموظف", "اسم المستفيد", "المستفيد",
        "اسم شاغل الوظيفة", "اسم شاغل/ة الوظيفة"
    ],
    "sector": [
        "القطاع التعليمي", "القطاع", "قطاع", "مكتب التعليم"
    ],
    "ministry_number": [
        "الرقم الوزاري", "رقم المدرسة الوزاري", "الرقم الوزاري للمدرسة",
        "رقم وزاري", "الرقم الوزارى"
    ],
    "school_name": [
        "اسم المدرسة", "المدرسة", "مدرسته", "المدرسة الحالية"
    ],
    "gender": [
        "الجنس", "النوع", "بنين بنات", "بنين/بنات", "الإدارة", "الادارة", "إدارة", "ادارة"
    ],
    "current_work": [
        "العمل الحالي", "العمل الحالى", "العمل", "نوع العمل",
        "المسمى الحالي", "المسمى الوظيفي", "الوظيفة", "نوع الوظيفة",
        "التكليف الحالي", "المكلف به"
    ],
    "category": [
        "الفئة", "الفئه", "فئة", "الفئه المستهدفة", "الفئة المستهدفة",
        "العمل الحالي", "العمل", "نوع العمل", "المسمى", "المسمى الوظيفي",
        "المسمى الحالي", "الوظيفة", "الوظيفه", "نوع الوظيفة", "نوع الوظيفه",
        "التكليف", "نوع التكليف", "فئة التكليف", "مجال التكليف",
        "التشكيل", "نوع التشكيل", "فئات التشكيلات", "فئة التشكيلات",
        "المكلف به", "التكليف الحالي", "الفئة الحالية", "الصفة"
    ],
    "raw_status": [
        "التجديد", "حالة التجديد", "الحالة", "الحلة", "حالة التكليف",
        "التوصية", "نوع التوصية", "القرار"
    ],
    "non_renewal_reason": [
        "المبرر في حال عدم التجديد", "المبرر فى حال عدم التجديد",
        "مبرر عدم التجديد", "المبرر", "سبب عدم التجديد", "السبب",
        "سبب عدم التجديد"
    ],
    "admin_note": [
        "ملاحظة", "ملاحظات", "الملاحظة", "الملاحظات",
        "ملاحظه", "ملاحظات إدارية", "ملاحظة إدارية",
        "ملحوظة", "ملحوظه", "الملحوظة", "الملحوظه"
    ],
    "mobile": [
        "الجوال", "رقم الجوال", "الهاتف", "رقم الهاتف", "جوال", "mobile", "phone"
    ],
}


def _normalized_header(value):
    text = normalize_text(value)
    text = text.replace("ـ", "")
    text = text.replace("إ", "ا").replace("أ", "ا").replace("آ", "ا")
    text = text.replace("ة", "ه")
    return text.strip().lower()


def _find_index(headers, aliases):
    normalized = [_normalized_header(h) for h in headers]
    wanted = [_normalized_header(a) for a in aliases]
    for alias in wanted:
        if alias in normalized:
            return normalized.index(alias)
    for index, header in enumerate(normalized):
        if any(alias and alias in header for alias in wanted):
            return index
    return None


def read_excel_records(file_path, preview_limit=25):
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    header_labels = {
        "national_id": "السجل المدني",
        "full_name": "الاسم",
        "raw_status": "التجديد",
        "school_name": "اسم المدرسة",
        "sector": "القطاع",
        "category": "العمل الحالي",
    }

    empty_result = {
        "rows": [],
        "preview_rows": [],
        "issue_rows": [],
        "total_rows": 0,
        "valid_count": 0,
        "invalid_count": 0,
        "duplicate_count": 0,
        "warning_count": 0,
        "category_counts": Counter(),
        "status_counts": Counter(),
        "category_status_counts": {},
        "category_summary": [],
        "status_summary": [],
        "issue_counts": Counter(),
        "missing_headers": [],
        "missing_header_labels": [],
        "recognized_headers": {},
        "has_mobile": False,
        "quality_level": "danger",
        "quality_label": "لا توجد بيانات",
        "quality_message": "لم يتم العثور على بيانات داخل الملف.",
        "can_confirm": False,
    }

    if not rows:
        return empty_result

    headers = [normalize_text(h) for h in rows[0]]
    indices = {key: _find_index(headers, aliases) for key, aliases in HEADER_ALIASES.items()}

    required = ["national_id", "full_name", "raw_status"]
    recommended = ["school_name", "sector", "category", "ministry_number", "gender", "non_renewal_reason", "admin_note"]

    missing_headers = [key for key in required if indices.get(key) is None]
    missing_header_labels = [header_labels.get(key, key) for key in missing_headers]

    recognized_headers = {}
    for key in required + recommended:
        index = indices.get(key)
        recognized_headers[key] = headers[index] if index is not None and index < len(headers) else ""

    parsed = []
    preview_rows = []
    issue_rows = []
    seen_ids = set()

    duplicate_count = 0
    invalid_count = 0
    warning_count = 0

    category_counts = Counter()
    status_counts = Counter()
    category_status_counts = {}
    issue_counts = Counter()
    has_mobile = indices.get("mobile") is not None

    for row_number, row in enumerate(rows[1:], start=2):
        if not row or not any(normalize_text(cell) for cell in row):
            continue

        def value_for(key):
            idx = indices.get(key)
            if idx is None or idx >= len(row):
                return ""
            return normalize_text(row[idx])

        national_id = normalize_national_id(value_for("national_id"))
        mobile_last4 = normalize_mobile_last4(value_for("mobile"))
        full_name = value_for("full_name")
        ministry_number = normalize_ministry_number(value_for("ministry_number"))
        school_name = value_for("school_name")
        sector = value_for("sector")
        gender = value_for("gender")

        current_work = value_for("current_work") or value_for("category")
        category = normalize_category(current_work or value_for("category"))

        if category == CATEGORY_UNKNOWN:
            row_text = " ".join(normalize_text(cell) for cell in row if normalize_text(cell))
            inferred_category = normalize_category(row_text)
            if inferred_category != CATEGORY_UNKNOWN:
                category = inferred_category

        raw_status = value_for("raw_status")
        status = classify_status(raw_status)
        category_key = category_group_key(category)

        non_renewal_reason = value_for("non_renewal_reason")
        admin_note = value_for("admin_note")

        errors = []
        warnings = []

        if not national_id:
            errors.append("السجل المدني مفقود")
            issue_counts["missing_national_id"] += 1

        if not full_name:
            errors.append("الاسم مفقود")
            issue_counts["missing_name"] += 1

        if not raw_status:
            errors.append("التجديد مفقود")
            issue_counts["missing_status"] += 1

        is_duplicate = bool(national_id and national_id in seen_ids)
        if is_duplicate:
            duplicate_count += 1
            issue_counts["duplicates"] += 1
            errors.append("سجل مدني مكرر داخل الملف")
        elif national_id:
            seen_ids.add(national_id)

        if category_key == "unknown":
            warnings.append("العمل الحالي غير مقروء")
            issue_counts["missing_category"] += 1

        if not school_name:
            warnings.append("اسم المدرسة غير متوفر")
            issue_counts["missing_school"] += 1

        if not sector:
            warnings.append("القطاع غير متوفر")
            issue_counts["missing_sector"] += 1

        if raw_status and status["key"] == "pending" and status["label"] == raw_status:
            warnings.append("حالة التجديد غير مصنفة آليًا")
            issue_counts["unknown_status"] += 1

        item = {
            "row_number": row_number,
            "national_id": national_id,
            "mobile_last4": mobile_last4,
            "full_name": full_name,
            "ministry_number": ministry_number,
            "school_name": school_name,
            "sector": sector,
            "gender": gender,
            "current_work": current_work or category,
            "category": category,
            "category_key": category_key,
            "raw_status": raw_status,
            "status_key": status["key"],
            "status_label": status["label"],
            "status_message": status["message"],
            "non_renewal_reason": non_renewal_reason,
            "admin_note": admin_note,
            "errors": errors,
            "warnings": warnings,
            "issues": errors + warnings,
            "is_valid": not errors,
        }

        if item["is_valid"]:
            parsed.append(item)
            category_counts[category_key] += 1
            status_counts[item["status_key"]] += 1
            category_status_counts.setdefault(category_key, Counter())
            category_status_counts[category_key][item["status_key"]] += 1
        else:
            invalid_count += 1

        if warnings:
            warning_count += 1

        if item["issues"] and len(issue_rows) < 30:
            issue_rows.append(item)

        if len(preview_rows) < preview_limit:
            preview_rows.append(item)

    category_summary = [
        {"key": "manager", "label": CATEGORY_MANAGER, "count": category_counts.get("manager", 0)},
        {"key": "deputy", "label": CATEGORY_DEPUTY, "count": category_counts.get("deputy", 0)},
        {"key": "counselor", "label": CATEGORY_COUNSELOR, "count": category_counts.get("counselor", 0)},
        {"key": "activity_leader", "label": CATEGORY_ACTIVITY, "count": category_counts.get("activity_leader", 0)},
        {"key": "unknown", "label": CATEGORY_UNKNOWN, "count": category_counts.get("unknown", 0)},
    ]

    status_summary = [
        {"key": "renewed", "label": "تم التجديد", "count": status_counts.get("renewed", 0)},
        {"key": "not_renewed", "label": "لم يتم التجديد", "count": status_counts.get("not_renewed", 0)},
        {"key": "limited", "label": "مدة محددة", "count": status_counts.get("limited", 0)},
        {"key": "pending", "label": "تحت الإجراء", "count": status_counts.get("pending", 0)},
    ]

    total_rows = len(parsed) + invalid_count

    if missing_headers or not parsed:
        quality_level = "danger"
        quality_label = "يتطلب معالجة"
        quality_message = "يوجد نقص في الأعمدة الأساسية أو لا توجد سجلات صحيحة للاعتماد."
    elif invalid_count:
        quality_level = "danger"
        quality_label = "يتطلب مراجعة"
        quality_message = "يوجد سجلات متجاوزة بسبب نقص بيانات أساسية أو تكرار داخل الملف."
    elif warning_count:
        quality_level = "warning"
        quality_label = "قابل للاعتماد مع ملاحظات"
        quality_message = "الملف قابل للاعتماد، مع وجود ملاحظات في بعض الحقول غير الأساسية."
    else:
        quality_level = "success"
        quality_label = "جاهز للاعتماد"
        quality_message = "الملف مكتمل ظاهريًا ولا توجد ملاحظات تمنع الاعتماد."

    return {
        "rows": parsed,
        "preview_rows": preview_rows,
        "issue_rows": issue_rows,
        "total_rows": total_rows,
        "valid_count": len(parsed),
        "invalid_count": invalid_count,
        "duplicate_count": duplicate_count,
        "warning_count": warning_count,
        "category_counts": category_counts,
        "status_counts": status_counts,
        "category_status_counts": category_status_counts,
        "category_summary": category_summary,
        "status_summary": status_summary,
        "issue_counts": issue_counts,
        "missing_headers": missing_headers,
        "missing_header_labels": missing_header_labels,
        "recognized_headers": recognized_headers,
        "has_mobile": has_mobile,
        "quality_level": quality_level,
        "quality_label": quality_label,
        "quality_message": quality_message,
        "can_confirm": bool(parsed and not missing_headers),
    }


def category_label_from_key(key):
    return CATEGORY_GROUPS.get(key, CATEGORY_UNKNOWN)


def status_label_from_key(key):
    return STATUS_LABELS.get(key, key or "")
